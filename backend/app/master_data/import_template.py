"""Imports a filled-in master-data workbook. Run via
`python -m scripts.import_master_data <file.xlsx>` (see backend/scripts/) —
dry run by default; pass --apply to actually write changes.

The whole run (every sheet) executes inside a single DB transaction: each row
is flushed (not committed) as it's processed, so later rows can resolve
foreign keys against rows this same run already inserted, and nothing is
committed unless --apply was passed AND zero row errors occurred across every
sheet — an all-or-nothing run, unlike the row-by-row-committed reference tool
this was modeled after.
"""

from __future__ import annotations

import argparse
import asyncio
import dataclasses
import datetime as dt
import uuid
from collections import defaultdict
from collections.abc import Callable
from decimal import Decimal
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import insert, select, update
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.db import AsyncSessionLocal, Base
from app.master_data.business_keys import BUSINESS_KEYS, GENERATED_CODE_COLUMN, KeyPart, Local, Ref
from app.master_data.enum_columns import ENUM_COLUMNS
from app.master_data.introspection import (
    SheetColumn,
    TableSpec,
    build_sheet_columns,
    build_sheet_titles,
    discover_tables,
    fk_target_table,
    humanize,
    resolve_key_parts,
)
from app.master_data.registry import CODE_GENERATOR_ENTITY
from app.services.code_generator import generate_code

Cache = dict[tuple[str, tuple[Any, ...]], uuid.UUID]


class RowValidationError(Exception):
    pass


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def _coerce(raw_value: Any, column) -> Any:
    if _is_blank(raw_value):
        return None
    py_type = column.type.python_type
    if py_type is bool:
        s = str(raw_value).strip().upper()
        if s in ("Y", "YES", "TRUE", "1"):
            return True
        if s in ("N", "NO", "FALSE", "0"):
            return False
        raise RowValidationError(f"expected Y or N, got {raw_value!r}")
    if py_type is int:
        return int(raw_value)
    if py_type is Decimal:
        return Decimal(str(raw_value))
    if py_type is dt.date:
        if isinstance(raw_value, dt.datetime):
            return raw_value.date()
        if isinstance(raw_value, dt.date):
            return raw_value
        return dt.datetime.strptime(str(raw_value).strip(), "%Y-%m-%d").date()
    if py_type is dt.datetime:
        if isinstance(raw_value, dt.datetime):
            return raw_value
        return dt.datetime.strptime(str(raw_value).strip(), "%Y-%m-%d %H:%M:%S")
    return str(raw_value).strip()


def _coerce_enum(raw_value: Any, enum_cls) -> str | None:
    if _is_blank(raw_value):
        return None
    s = str(raw_value).strip()
    for member in enum_cls:
        if member.value.lower() == s.lower():
            return member.value
    raise RowValidationError(f"must be one of: {', '.join(m.value for m in enum_cls)} (got {raw_value!r})")


async def _query_existing(session: AsyncSession, table_name: str, parts: tuple[KeyPart, ...], local_key: list[Any]) -> uuid.UUID | None:
    table = Base.metadata.tables[table_name]
    conditions = []
    for part, value in zip(parts, local_key, strict=True):
        col_name = part.column if isinstance(part, Local) else part.fk_column
        conditions.append(table.c[col_name] == value)
    stmt = select(table.c.id).where(*conditions)
    rows = (await session.execute(stmt)).scalars().all()
    if len(rows) > 1:
        raise RowValidationError(f"multiple existing rows in '{table_name}' match this key — the key is not actually unique in the data")
    return rows[0] if rows else None


async def resolve_id(session: AsyncSession, cache: Cache, target_table: str, flat_values: list[Any]) -> uuid.UUID | None:
    """flat_values must be exactly len(resolve_key_parts(target_table)) raw
    (uncoerced) cell values, in that same order — i.e. exactly what's under
    an FK's expanded sheet columns for one row."""
    parts = BUSINESS_KEYS.get(target_table, ())
    if not parts:
        return None
    local_key: list[Any] = []
    pos = 0
    for part in parts:
        if isinstance(part, Local):
            column = Base.metadata.tables[target_table].columns[part.column]
            local_key.append(_coerce(flat_values[pos], column))
            pos += 1
        else:
            nested_target = fk_target_table(target_table, part.fk_column)
            n = len(resolve_key_parts(nested_target))
            nested_id = await resolve_id(session, cache, nested_target, flat_values[pos : pos + n])
            pos += n
            if nested_id is None:
                return None
            local_key.append(nested_id)
    if any(v is None for v in local_key):
        return None
    cache_key = (target_table, tuple(local_key))
    if cache_key in cache:
        return cache[cache_key]
    row_id = await _query_existing(session, target_table, parts, local_key)
    if row_id is not None:
        cache[cache_key] = row_id
    return row_id


async def _import_row(session: AsyncSession, spec: TableSpec, columns: list[SheetColumn], raw: dict[SheetColumn, Any], cache: Cache) -> str:
    raw_by_plain: dict[str, Any] = {}
    raw_by_fk: dict[str, list[Any]] = defaultdict(list)
    for sc in columns:
        if sc.plain is not None:
            raw_by_plain[sc.plain.name] = raw[sc]
        elif sc.fk is not None:
            raw_by_fk[sc.fk.fk_column].append(raw[sc])

    resolved_plain: dict[str, Any] = {}
    for col in spec.plain_columns:
        raw_value = raw_by_plain.get(col.name)
        enum_cls = ENUM_COLUMNS.get(spec.name, {}).get(col.name)
        try:
            value = _coerce_enum(raw_value, enum_cls) if enum_cls is not None else _coerce(raw_value, col.column)
        except RowValidationError as exc:
            raise RowValidationError(f"{humanize(col.name)}: {exc}") from exc
        is_generated_code = spec.generated_code_column == col.name
        if value is None and col.required and not is_generated_code:
            raise RowValidationError(f"{humanize(col.name)} is required")
        resolved_plain[col.name] = value

    resolved_fk: dict[str, uuid.UUID | None] = {}
    for fk in spec.fk_expansions:
        flat = raw_by_fk.get(fk.fk_column, [])
        if all(_is_blank(v) for v in flat):
            if fk.required:
                raise RowValidationError(f"{humanize(fk.fk_column)} is required")
            resolved_fk[fk.fk_column] = None
            continue
        target_id = await resolve_id(session, cache, fk.target_table, flat)
        if target_id is None:
            labels = ", ".join(p.label for p in fk.key_parts)
            raise RowValidationError(f"{humanize(fk.fk_column)}: no matching row in '{fk.target_table}' for {labels} = {tuple(flat)!r}")
        resolved_fk[fk.fk_column] = target_id

    values: dict[str, Any] = {**resolved_plain, **resolved_fk}

    business_parts = BUSINESS_KEYS.get(spec.name, ())
    is_blank_generated_code = spec.generated_code_column is not None and _is_blank(raw_by_plain.get(spec.generated_code_column))

    existing_id: uuid.UUID | None = None
    own_key_tuple: tuple[Any, ...] | None = None
    if business_parts and not is_blank_generated_code:
        own_key_tuple = tuple(
            resolved_plain[p.column] if isinstance(p, Local) else resolved_fk[p.fk_column] for p in business_parts
        )
        if any(v is None for v in own_key_tuple):
            raise RowValidationError("this table's lookup key columns cannot be blank")
        cache_key = (spec.name, own_key_tuple)
        existing_id = cache.get(cache_key)
        if existing_id is None:
            existing_id = await _query_existing(session, spec.name, business_parts, list(own_key_tuple))

    table = Base.metadata.tables[spec.name]
    now = dt.datetime.now(dt.UTC)

    if existing_id is not None:
        update_values = dict(values)
        if "updated_at" in table.columns:
            update_values["updated_at"] = now
        if update_values:
            await session.execute(update(table).where(table.c.id == existing_id).values(**update_values))
        row_id = existing_id
        outcome = "updated"
    else:
        if is_blank_generated_code:
            entity_code = CODE_GENERATOR_ENTITY[spec.name]
            code_col = spec.generated_code_column
            assert code_col is not None
            values[code_col] = await generate_code(session, entity_code)
            if business_parts:
                own_key_tuple = tuple(
                    values[p.column] if isinstance(p, Local) else resolved_fk[p.fk_column] for p in business_parts
                )
        new_id = uuid.uuid4()
        insert_values: dict[str, Any] = {"id": new_id, **values}
        if "created_at" in table.columns:
            insert_values["created_at"] = now
        if "updated_at" in table.columns:
            insert_values["updated_at"] = now
        await session.execute(insert(table).values(**insert_values))
        row_id = new_id
        outcome = "inserted"

    await session.flush()
    if business_parts and own_key_tuple is not None:
        cache[(spec.name, own_key_tuple)] = row_id
    return outcome


@dataclasses.dataclass
class RowError:
    sheet: str
    row: int
    message: str


@dataclasses.dataclass
class SheetStats:
    inserted: int = 0
    updated: int = 0


def _header_text(raw: Any) -> str:
    s = str(raw).strip()
    return s[:-2] if s.endswith(" *") else s


def _match_headers(ws: Worksheet, columns: list[SheetColumn], sheet_name: str, errors: list[RowError]) -> dict[SheetColumn, int] | None:
    by_text: dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        raw = ws.cell(row=1, column=col_idx).value
        if raw:
            by_text[_header_text(raw)] = col_idx

    mapping: dict[SheetColumn, int] = {}
    missing: list[str] = []
    for col in columns:
        idx = by_text.get(col.header)
        if idx is None:
            if col.required:
                missing.append(col.header)
            continue
        mapping[col] = idx
    if missing:
        errors.append(RowError(sheet_name, 1, f"Missing required column(s): {', '.join(missing)}"))
        return None
    return mapping


async def _import_sheet(session: AsyncSession, spec: TableSpec, ws: Worksheet, cache: Cache, stats: dict[str, SheetStats], errors: list[RowError]) -> None:
    columns = build_sheet_columns(spec)
    header_map = _match_headers(ws, columns, spec.name, errors)
    if header_map is None:
        return
    stat = stats.setdefault(spec.name, SheetStats())
    for row_idx in range(2, ws.max_row + 1):
        raw = {col: ws.cell(row=row_idx, column=idx).value for col, idx in header_map.items()}
        if all(_is_blank(v) for v in raw.values()):
            continue
        try:
            outcome = await _import_row(session, spec, columns, raw, cache)
        except RowValidationError as exc:
            errors.append(RowError(spec.name, row_idx, str(exc)))
            continue
        if outcome == "inserted":
            stat.inserted += 1
        else:
            stat.updated += 1


@dataclasses.dataclass
class ImportResult:
    stats: dict[str, SheetStats]
    errors: list[RowError]
    applied: bool


async def run_import(
    path: str,
    apply: bool,
    only: set[str] | None,
    session_factory: Callable[[], AsyncSession] = AsyncSessionLocal,
) -> ImportResult:
    table_specs = discover_tables()
    sheet_titles = build_sheet_titles(table_specs)
    wb = load_workbook(path, data_only=True)

    cache: Cache = {}
    stats: dict[str, SheetStats] = {}
    errors: list[RowError] = []

    async with session_factory() as session:
        for spec in table_specs:
            title = sheet_titles[spec.name]
            if title not in wb.sheetnames:
                continue
            if only is not None and spec.name.lower() not in only and title.lower() not in only:
                continue
            await _import_sheet(session, spec, wb[title], cache, stats, errors)

        if apply and not errors:
            await session.commit()
        else:
            await session.rollback()

    return ImportResult(stats=stats, errors=errors, applied=apply and not errors)


def _print_report(result: ImportResult) -> None:
    print("\nSummary:")
    for sheet, stat in result.stats.items():
        print(f"  {sheet}: {stat.inserted} inserted, {stat.updated} updated")

    if result.errors:
        by_sheet: dict[str, list[RowError]] = defaultdict(list)
        for err in result.errors:
            by_sheet[err.sheet].append(err)
        print(f"\n{len(result.errors)} row error(s):")
        for sheet, errs in by_sheet.items():
            print(f"  {sheet}:")
            for err in errs[:15]:
                print(f"    row {err.row}: {err.message}")
            if len(errs) > 15:
                print(f"    ...and {len(errs) - 15} more")

    print()
    if result.applied:
        print("Applied.")
    elif result.errors:
        print("NOT applied — fix the errors above and re-run.")
    else:
        print("Dry run — no changes written. Re-run with --apply to write them.")


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(description="Import a filled-in master-data workbook.")
    parser.add_argument("path", help="Path to the filled .xlsx file")
    parser.add_argument("--apply", action="store_true", help="Actually write changes (default is dry run)")
    parser.add_argument("--sheets", default=None, help="Comma-separated list of sheet/table names to limit to")
    args = parser.parse_args(argv)

    only = {s.strip().lower() for s in args.sheets.split(",")} if args.sheets else None
    result = asyncio.run(run_import(args.path, args.apply, only))
    _print_report(result)
    if result.errors:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
