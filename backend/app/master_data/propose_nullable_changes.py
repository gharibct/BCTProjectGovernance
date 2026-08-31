"""Turns 'the user hid a currently-required column in a filled-in
master-data workbook' into a reviewable proposal to make that column
nullable: the exact `ALTER TABLE ... DROP NOT NULL` plus a before/after
snippet of the matching db/tables/*.sql line and the matching ORM model
line, written to a timestamped Markdown file. Never edits db/tables/*.sql,
app/models/*.py, or the live DB — read and propose only.

Run via `python -m scripts.propose_nullable_changes <file.xlsx>` (see
backend/scripts/). Only plain (non-foreign-key) required columns are
proposable in this first version — hiding a required foreign-key column
isn't turned into a proposal.

The matching DDL/model file is found by searching db/tables/*.sql and
app/models/*.py for the table's own CREATE TABLE / __tablename__, rather
than a hand-maintained table -> file lookup, so it can't go stale as those
directories grow.
"""

from __future__ import annotations

import argparse
import datetime as dt
import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.master_data.introspection import TableSpec, build_sheet_columns, build_sheet_titles, discover_tables

BACKEND_DIR = Path(__file__).resolve().parents[2]
REPO_ROOT = BACKEND_DIR.parent
SQL_TABLES_DIR = REPO_ROOT / "db" / "tables"
MODELS_DIR = BACKEND_DIR / "app" / "models"


def _find_sql_definition(table: str, column: str) -> tuple[Path | None, str | None]:
    column_re = re.compile(rf"^\s*{re.escape(column)}\b", re.IGNORECASE)
    for sql_file in sorted(SQL_TABLES_DIR.glob("*.sql")):
        text = sql_file.read_text(encoding="utf-8")
        if re.search(rf"CREATE TABLE\s+{re.escape(table)}\b", text, re.IGNORECASE) is None:
            continue
        for line in text.splitlines():
            if column_re.match(line):
                return sql_file, line.strip()
        return sql_file, None
    return None, None


def _find_model_definition(table: str, column: str) -> tuple[Path | None, str | None]:
    column_re = re.compile(rf"^\s*{re.escape(column)}\s*:")
    for model_file in sorted(MODELS_DIR.glob("*.py")):
        text = model_file.read_text(encoding="utf-8")
        if re.search(rf'__tablename__\s*=\s*"{re.escape(table)}"', text) is None:
            continue
        for line in text.splitlines():
            if column_re.match(line):
                return model_file, line.strip()
        return model_file, None
    return None, None


def find_hidden_required_columns(path: str) -> list[tuple[TableSpec, str]]:
    """(TableSpec, column_name) pairs where the sheet column is currently
    required (starred, plain DB column) per the live schema, but the user
    has hidden it in the given workbook."""
    table_specs = discover_tables()
    sheet_titles = build_sheet_titles(table_specs)
    wb = load_workbook(path)

    found: list[tuple[TableSpec, str]] = []
    for spec in table_specs:
        title = sheet_titles[spec.name]
        if title not in wb.sheetnames:
            continue
        ws = wb[title]
        by_text: dict[str, int] = {}
        for col_idx in range(1, ws.max_column + 1):
            raw = ws.cell(row=1, column=col_idx).value
            if raw:
                text = str(raw).strip()
                by_text[text[:-2] if text.endswith(" *") else text] = col_idx

        for col in build_sheet_columns(spec):
            if not col.required or col.plain is None:
                continue  # FK-expansion columns aren't proposable in this version
            col_idx = by_text.get(col.header)
            if col_idx is None:
                continue
            if ws.column_dimensions[get_column_letter(col_idx)].hidden:
                found.append((spec, col.plain.name))
    return found


def write_proposal(source_path: str, found: list[tuple[TableSpec, str]]) -> Path:
    timestamp = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    out_path = BACKEND_DIR / f"master_data_nullable_proposal_{timestamp}.md"

    lines = [
        f"# Nullable-column proposal — generated from `{source_path}`",
        "",
        "Hiding a required (starred) column in the master-data workbook proposes making it",
        "nullable. Nothing has been changed by this script — review each item below and apply",
        "by hand if agreed.",
        "",
    ]
    if not found:
        lines.append("No hidden required columns found — nothing to propose.")

    for spec, column in found:
        sql_file, sql_line = _find_sql_definition(spec.name, column)
        model_file, model_line = _find_model_definition(spec.name, column)

        lines.append(f"## {spec.name}.{column}")
        lines.append("")
        lines.append("```sql")
        lines.append(f"ALTER TABLE {spec.name} ALTER COLUMN {column} DROP NOT NULL;")
        lines.append("```")
        lines.append("")

        if sql_file is not None:
            lines.append(f"DDL file: `{sql_file.relative_to(REPO_ROOT)}`")
            if sql_line:
                lines.append(f"- Current:  `{sql_line}`")
                lines.append(f"- Proposed: `{sql_line.replace(' NOT NULL', '')}`")
        else:
            lines.append(f"DDL file: not found automatically — search `db/tables/` for `CREATE TABLE {spec.name}` by hand.")
        lines.append("")

        if model_file is not None:
            lines.append(f"Model file: `{model_file.relative_to(REPO_ROOT)}`")
            if model_line:
                proposed = re.sub(r"Mapped\[([^|\]]+)\]", r"Mapped[\1 | None]", model_line, count=1)
                lines.append(f"- Current:  `{model_line}`")
                lines.append(f"- Proposed: `{proposed}`")
        else:
            lines.append(f'Model file: not found automatically — search `app/models/` for `__tablename__ = "{spec.name}"` by hand.')
        lines.append("")

    out_path.write_text("\n".join(lines), encoding="utf-8")
    return out_path


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(description="Propose nullable-column changes from a filled-in workbook's hidden columns.")
    parser.add_argument("path", help="Path to the filled .xlsx file")
    args = parser.parse_args(argv)

    found = find_hidden_required_columns(args.path)
    out_path = write_proposal(args.path, found)
    print(f"Wrote {len(found)} proposal(s) to {out_path}")


if __name__ == "__main__":
    main()
