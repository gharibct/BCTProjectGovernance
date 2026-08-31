"""Builds the multi-sheet master-data Excel workbook. Run via
`python -m scripts.generate_master_data_template` (see backend/scripts/).
"""

from __future__ import annotations

import argparse

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from app.master_data.business_keys import GENERATED_CODE_COLUMN
from app.master_data.introspection import TableSpec, build_sheet_columns, build_sheet_titles, discover_tables, used_lists

MAX_VALIDATION_ROWS = 500
DEFAULT_OUTPUT = "Master-Data-Template.xlsx"


def build_workbook(table_specs: list[TableSpec]) -> Workbook:
    wb = Workbook()
    _write_readme(wb, table_specs)
    ranges = _write_lists_sheet(wb, used_lists(table_specs))
    sheet_titles = build_sheet_titles(table_specs)
    for spec in table_specs:
        _write_table_sheet(wb, spec, ranges, sheet_titles[spec.name])
    return wb


def _write_readme(wb: Workbook, table_specs: list[TableSpec]) -> None:
    ws = wb.active
    ws.title = "Read Me"
    ws.column_dimensions["A"].width = 110
    generated_tables = ", ".join(sorted(GENERATED_CODE_COLUMN))
    no_key_tables = ", ".join(sorted(s.name for s in table_specs if not s.business_key))
    lines = [
        "Master Data Template",
        "",
        "One sheet per database table. Read this sheet fully before filling in data.",
        "",
        "REQUIRED COLUMNS are marked with a trailing ' *' in the header and are always visible.",
        "OPTIONAL COLUMNS that are already nullable today are hidden by default — unhide a column",
        "(select the columns on either side, right-click, Unhide) if you want to fill it in.",
        "",
        "HIDING A REQUIRED COLUMN is a proposal, not a live change: if you hide a currently",
        "required (starred) column and run `python -m scripts.propose_nullable_changes <file>`,",
        "it writes out the exact database and model changes needed to make that column optional,",
        "for someone to review and apply by hand. Nothing is changed automatically by that script.",
        "",
        "FOREIGN KEY columns are entered as a human-readable value from the referenced sheet",
        "(e.g. a Project Code, a user's Email) — never as a raw id. Hover a column header (the",
        "little red-cornered cells) for exactly which sheet and column to use.",
        "",
        "BUSINESS-KEY / CODE columns identify a row for import:",
        "  - Leave it blank on a row for one of these tables to have a new code generated",
        f"    automatically: {generated_tables}.",
        "  - Fill in an existing code/key to update that row instead of inserting a new one.",
        "  - Fill in a code/key that doesn't exist yet to insert a new row using that exact value",
        "    (useful for migrating legacy codes from another system).",
        "  - These tables have no business key at all — every filled-in row is always inserted",
        f"    as new: {no_key_tables}.",
        "",
        "BOOLEAN columns: enter Y or N.",
        "DATE columns: enter as YYYY-MM-DD.",
        "",
        "Import with: python -m scripts.import_master_data <this file>",
        "  (dry run by default — add --apply to actually write changes;",
        "   add --sheets=SheetA,SheetB to limit the run to specific sheets).",
    ]
    for i, line in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True, size=14)


def _write_lists_sheet(wb: Workbook, used: dict[str, list[str]]) -> dict[str, str]:
    """A hidden sheet holding one column per distinct enum (plus Y/N), used
    as the source range for dropdown data validation — a worksheet range
    avoids Excel's 255-character limit on an inline validation list, which
    some of this schema's longer enums (e.g. ApplicablePhase) would exceed."""
    ws = wb.create_sheet(title="Lists")
    ws.sheet_state = "hidden"
    ranges: dict[str, str] = {}
    for idx, (name, values) in enumerate(used.items(), start=1):
        col_letter = get_column_letter(idx)
        ws.cell(row=1, column=idx, value=name)
        for r, v in enumerate(values, start=2):
            ws.cell(row=r, column=idx, value=v)
        ranges[name] = f"Lists!${col_letter}$2:${col_letter}${1 + len(values)}"
    return ranges


def _write_table_sheet(wb: Workbook, spec: TableSpec, ranges: dict[str, str], title: str) -> None:
    ws: Worksheet = wb.create_sheet(title=title)
    columns = build_sheet_columns(spec)

    for idx, col in enumerate(columns, start=1):
        header = col.header + (" *" if col.required else "")
        cell = ws.cell(row=1, column=idx, value=header)
        cell.font = Font(bold=True)
        if col.note:
            cell.comment = Comment(col.note, "Master Data Template")
        col_letter = get_column_letter(idx)
        ws.column_dimensions[col_letter].width = max(14, len(header) + 4)
        if not col.required:
            ws.column_dimensions[col_letter].hidden = True
        if col.list_key:
            dv = DataValidation(
                type="list",
                formula1=f"={ranges[col.list_key]}",
                allow_blank=True,
                showErrorMessage=True,
                error="Value must come from the dropdown list — see the column note.",
            )
            ws.add_data_validation(dv)
            dv.add(f"{col_letter}2:{col_letter}{MAX_VALIDATION_ROWS + 1}")

    ws.freeze_panes = "A2"


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(description="Generate the master-data Excel template.")
    parser.add_argument("--out", default=DEFAULT_OUTPUT, help="Output .xlsx path")
    args = parser.parse_args(argv)

    table_specs = discover_tables()
    wb = build_workbook(table_specs)
    wb.save(args.out)
    print(f"Wrote {len(table_specs)} table sheets to {args.out}")


if __name__ == "__main__":
    main()
