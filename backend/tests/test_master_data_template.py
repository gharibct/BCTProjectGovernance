from openpyxl.utils import get_column_letter

from app.master_data.generate_template import build_workbook
from app.master_data.introspection import build_sheet_columns, build_sheet_titles, discover_tables


def test_discover_tables_excludes_system_tables():
    names = {s.name for s in discover_tables()}
    for excluded in ("user_activity_log", "ai_field_suggestions", "ai_row_suggestions", "backup_restore_log", "id_sequences", "project_documents"):
        assert excluded not in names
    assert "risk_log" in names
    assert "projects" in names


def test_computed_columns_excluded():
    specs = discover_tables()
    projects = next(s for s in specs if s.name == "projects")
    plain_names = {c.name for c in projects.plain_columns}
    assert "planned_duration_days" not in plain_names
    assert "actual_duration_days" not in plain_names
    assert "delivery_declared_overall_health" not in plain_names
    assert "id" not in plain_names
    assert "created_at" not in plain_names


def test_required_marking_matches_db_nullability():
    specs = discover_tables()
    risk = next(s for s in specs if s.name == "risk_log")
    by_header = {c.header: c for c in build_sheet_columns(risk)}
    assert by_header["Risk Title"].required is True
    assert by_header["Escalated To"].required is False


def test_fk_column_shows_business_key_not_raw_id():
    specs = discover_tables()
    risk = next(s for s in specs if s.name == "risk_log")
    headers = [c.header for c in build_sheet_columns(risk)]
    assert "Project" in headers
    assert "Project Id" not in headers
    assert "Project ID" not in headers


def test_composite_fk_expands_into_labelled_parts():
    specs = discover_tables()
    defects = next(s for s in specs if s.name == "measurement_development_defects")
    headers = {c.header for c in build_sheet_columns(defects)}
    assert "Measurement - Project Code" in headers
    assert "Measurement - Reporting Period Code" in headers


def test_workbook_has_one_sheet_per_included_table_plus_readme_and_lists():
    specs = discover_tables()
    wb = build_workbook(specs)
    titles = build_sheet_titles(specs)
    assert "Read Me" in wb.sheetnames
    assert "Lists" in wb.sheetnames
    for spec in specs:
        assert titles[spec.name] in wb.sheetnames
    assert len(wb.sheetnames) == len(specs) + 2


def test_nullable_column_hidden_required_column_visible():
    specs = discover_tables()
    wb = build_workbook(specs)
    titles = build_sheet_titles(specs)
    ws = wb[titles["risk_log"]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    title_letter = get_column_letter(headers.index("Risk Title *") + 1)
    desc_letter = get_column_letter(headers.index("Risk Description") + 1)

    assert ws.column_dimensions[title_letter].hidden is False
    assert ws.column_dimensions[desc_letter].hidden is True


def test_enum_column_gets_dropdown_and_note():
    specs = discover_tables()
    risk = next(s for s in specs if s.name == "risk_log")
    columns = build_sheet_columns(risk)
    status = next(c for c in columns if c.header == "Current Status")
    assert status.list_key == "RiskStatus"
    assert "Open" in status.note
